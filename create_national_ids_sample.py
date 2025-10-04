#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ایجاد فایل نمونه اکسل برای کدهای ملی مجاز
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_sample_national_ids_file():
    """ایجاد فایل نمونه کدهای ملی مجاز"""
    
    # ایجاد workbook جدید
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "کدهای ملی مجاز"
    
    # تنظیم عرض ستون‌ها
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    
    # استایل‌های مورد نیاز
    header_font = Font(name='B Nazanin', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='B Nazanin', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # هدر جدول
    headers = ['کد ملی', 'نام دانش آموز', 'کلاس']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # داده‌های نمونه
    sample_data = [
        ['1234567890', 'علی احمدی', '۷۰۱'],
        ['0987654321', 'فاطمه رضایی', '۷۰۲'],
        ['1122334455', 'محمد علیزاده', '۷۰۳'],
        ['5544332211', 'زهرا محمدی', '۸۰۱'],
        ['6677889900', 'حسین کریمی', '۸۰۲'],
        ['9988776655', 'مریم حسینی', '۸۰۳'],
        ['1357924680', 'رضا نوری', '۸۰۴'],
        ['2468013579', 'سارا احمدی', '۸۰۵'],
        ['3691470258', 'امیر رضایی', '۸۰۶'],
        ['7410852963', 'نرگس کاظمی', '۸۰۷'],
    ]
    
    # اضافه کردن داده‌ها
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
    
    # اضافه کردن یک sheet راهنما
    ws_guide = wb.create_sheet("راهنما")
    ws_guide.column_dimensions['A'].width = 80
    
    guide_text = [
        "📋 راهنمای استفاده از فایل کدهای ملی مجاز",
        "",
        "🔹 این فایل برای تعریف کدهای ملی مجاز برای ثبت نام استفاده می‌شود.",
        "",
        "📝 ساختار فایل:",
        "   • ستون A: کد ملی (10 رقم)",
        "   • ستون B: نام دانش آموز (اختیاری)",
        "   • ستون C: کلاس (اختیاری)",
        "",
        "⚠️ نکات مهم:",
        "   1. کد ملی باید دقیقاً 10 رقم باشد",
        "   2. کد ملی باید معتبر باشد",
        "   3. هر کد ملی فقط یک بار قابل استفاده است",
        "   4. نام و کلاس فقط برای مرجع است و اجباری نیست",
        "   5. سطر اول (هدر) را حذف نکنید",
        "",
        "📤 نحوه استفاده:",
        "   1. این فایل را با اطلاعات دانش آموزان خود پر کنید",
        "   2. فایل را ذخیره کنید",
        "   3. وارد پنل مدیریت شوید",
        "   4. روی دکمه 'آپلود کدهای ملی مجاز' کلیک کنید",
        "   5. این فایل را انتخاب کنید",
        "",
        "✅ بعد از آپلود:",
        "   • فقط دانش آموزانی که کد ملی‌شان در این لیست است می‌توانند ثبت نام کنند",
        "   • هر کد ملی فقط یک بار قابل استفاده است",
        "   • می‌توانید لیست را به‌روزرسانی کنید",
        "",
        "🔍 مثال:",
        "   کد ملی: 1234567890",
        "   نام: علی احمدی",
        "   کلاس: ۷۰۱",
        "",
        f"📅 تاریخ ایجاد: {datetime.now().strftime('%Y/%m/%d')}",
        "🏫 مدرسه هیات امنایی نخبگان",
    ]
    
    for row_idx, text in enumerate(guide_text, start=1):
        cell = ws_guide.cell(row=row_idx, column=1)
        cell.value = text
        cell.font = Font(name='B Nazanin', size=11)
        cell.alignment = Alignment(horizontal='right', vertical='top', wrap_text=True)
    
    # ذخیره فایل
    filename = 'sample_national_ids.xlsx'
    wb.save(filename)
    print(f"✅ فایل نمونه با موفقیت ایجاد شد: {filename}")
    print(f"📁 مسیر: {filename}")
    print("\n📋 محتویات فایل:")
    print("   • 10 کد ملی نمونه")
    print("   • یک sheet راهنما")
    print("\n🎯 این فایل را می‌توانید:")
    print("   1. به عنوان نمونه استفاده کنید")
    print("   2. با اطلاعات واقعی پر کنید")
    print("   3. در سیستم آپلود کنید")

if __name__ == "__main__":
    try:
        create_sample_national_ids_file()
    except ImportError:
        print("❌ خطا: کتابخانه openpyxl نصب نیست")
        print("💡 برای نصب: pip install openpyxl")
    except Exception as e:
        print(f"❌ خطا: {e}")
